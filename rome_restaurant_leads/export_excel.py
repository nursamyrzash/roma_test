from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT = Path("data/output")


def _usage_dataframe(usage):
    d = usage.as_dict()
    rows = [
        ["Run started (UTC)", d.get("run_started_at", ""), ""],
        ["Run finished (UTC)", d.get("run_finished_at", ""), ""],
        ["Text Search API calls", d.get("text_search_calls", 0), "Text Search Essentials (IDs Only) field mask"],
        ["Text Search pages received", d.get("text_search_pages", 0), "Up to SEARCH_MAX_PAGES per query/grid cell"],
        ["Place Details API calls", d.get("place_details_calls", 0), "One per unique Place ID"],
        ["Total Google API calls", d.get("total_google_calls", 0), "Search + Details"],
        ["Google API errors", d.get("google_api_errors", 0), "HTTP/API failures"],
        ["Raw places returned", d.get("raw_places_returned", 0), "Before Place ID deduplication"],
        ["Unique places found", d.get("unique_places_found", 0), "After Place ID deduplication"],
        ["Duplicates removed", d.get("duplicate_places_removed", 0), "Raw - unique"],
        ["Google calls / unique place", d.get("google_calls_per_unique_place", 0), "Efficiency metric"],
        ["Website rows checked", d.get("websites_checked", 0), "No Google API charge"],
        ["Rows with website URL", d.get("websites_with_url", 0), ""],
        ["Rows without website URL", d.get("websites_without_url", 0), ""],
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value", "Notes"])


def export_excel(rows, usage):
    OUTPUT.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)

    all_cols = [
        "place_id", "restaurant_name", "address", "google_maps_url",
        "latitude", "longitude", "primary_type", "types", "business_status",
        "phone", "rating", "review_count", "price_level", "website_url",
        "website_exists", "website_status", "https", "mobile_ready",
        "response_time_ms", "english_version", "menu_online", "menu_pdf_only",
        "online_booking", "whatsapp", "google_maps_on_site", "instagram",
        "site_title", "pages_checked", "site_error",
        "lead_score", "lead_priority", "reason",
        "contact_status", "contact_date", "response", "follow_up_date",
        "client_status"
    ]
    for c in all_cols:
        if c not in df.columns:
            df[c] = ""

    df = df[all_cols].sort_values(
        ["lead_score", "review_count"], ascending=[False, False]
    )

    hot = df[df["lead_score"] >= 80].copy()
    warm = df[(df["lead_score"] >= 60) & (df["lead_score"] < 80)].copy()
    audit_cols = [
        "restaurant_name", "website_url", "website_status", "https",
        "mobile_ready", "response_time_ms", "english_version", "menu_online",
        "menu_pdf_only", "online_booking", "whatsapp", "google_maps_on_site",
        "instagram", "site_title", "pages_checked", "site_error", "lead_score"
    ]
    audit = df[audit_cols].copy()
    usage_df = _usage_dataframe(usage)

    path = OUTPUT / "rome_restaurant_leads.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="ALL_RESTAURANTS", index=False)
        hot.to_excel(writer, sheet_name="HOT_LEADS", index=False)
        warm.to_excel(writer, sheet_name="WARM_LEADS", index=False)
        audit.to_excel(writer, sheet_name="SITE_AUDIT", index=False)
        usage_df.to_excel(writer, sheet_name="API_USAGE", index=False)

        config = pd.DataFrame([
            ["Score 80-100", "HOT", "Contact first"],
            ["Score 60-79", "WARM", "Contact after HOT"],
            ["Score 40-59", "MEDIUM", "Optional"],
            ["Score 0-39", "LOW", "Low priority"],
        ], columns=["Score", "Priority", "Action"])
        config.to_excel(writer, sheet_name="CONFIG", index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = 0
            for cell in col[:200]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 45)

    wb.save(path)
    return path
